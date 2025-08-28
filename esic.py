import streamlit as st
import pandas as pd
import re
import io
import zipfile
from datetime import datetime
import logging
import traceback
from pathlib import Path
from io import BytesIO

# PDF processing libraries
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# UTILITY FUNCTIONS FOR NUMERIC CONVERSION
# ============================================================================

def safe_numeric_convert(value, is_integer=False):
    """Safely convert string to number, removing commas and handling decimals"""
    try:
        if not value or value == '-' or str(value).lower() in ['not found', 'n/a', 'error', '']:
            return 0 if is_integer else 0.0
        
        # Remove commas and clean the string
        clean_value = str(value).replace(',', '').strip()
        
        if is_integer:
            return int(float(clean_value))
        else:
            return float(clean_value)
    except (ValueError, TypeError):
        return 0 if is_integer else 0.0

def safe_numeric_convert_challan(value, is_integer=False):
    """Safely convert string to number for challan data"""
    try:
        if not value or str(value).lower() in ['not found', 'n/a', 'error', '']:
            return 0 if is_integer else 0.0
        
        # Remove currency symbols, commas, and clean the string
        clean_value = str(value).replace('₹', '').replace(',', '').strip()
        
        if is_integer:
            return int(float(clean_value))
        else:
            return float(clean_value)
    except (ValueError, TypeError):
        return value  # Return original if conversion fails

def extract_month_from_text(text):
    """Extract month name from the contribution history line"""
    try:
        # Look for patterns like "for Apr2024", "for Jan2024", etc.
        month_pattern = r'for\s+([A-Za-z]{3,9}\d{4})'
        match = re.search(month_pattern, text)
        
        if match:
            period_str = match.group(1)
            
            # Extract month name (first 3+ letters before the year)
            month_match = re.match(r'([A-Za-z]{3,9})', period_str)
            if month_match:
                month_name = month_match.group(1).capitalize()
                
                # Map common abbreviations to full month names
                month_mapping = {
                    'Jan': 'January', 'Feb': 'February', 'Mar': 'March',
                    'Apr': 'April', 'May': 'May', 'Jun': 'June',
                    'Jul': 'July', 'Aug': 'August', 'Sep': 'September',
                    'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
                }
                
                return month_mapping.get(month_name, month_name)
        
        # Alternative patterns
        patterns = [
            r'Contribution\s+History.*?for\s+([A-Za-z]+)',
            r'ECR\s+Of.*?for\s+([A-Za-z]+)',
            r'Period[:\s]+([A-Za-z]+\s*\d{4})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                month_part = match.group(1).strip()
                # Extract just the alphabetic part
                month_alpha = re.match(r'([A-Za-z]+)', month_part)
                if month_alpha:
                    month_name = month_alpha.group(1).capitalize()
                    month_mapping = {
                        'Jan': 'January', 'Feb': 'February', 'Mar': 'March',
                        'Apr': 'April', 'May': 'May', 'Jun': 'June',
                        'Jul': 'July', 'Aug': 'August', 'Sep': 'September',
                        'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
                    }
                    return month_mapping.get(month_name, month_name)
        
        return "Not Found"
        
    except Exception as e:
        logger.error(f"Error extracting month: {str(e)}")
        return "Not Found"

# ============================================================================
# ESIC CONTRIBUTION HISTORY EXTRACTOR
# ============================================================================

def extract_esic_data(pdf_file):
    """Extract ESIC ecr data from PDF while preserving structure"""
    try:    
        with pdfplumber.open(pdf_file) as pdf:
            extracted_data = {
                'header_info': {},
                'summary_info': {},
                'employee_data': [],
                'footer_info': {}
            }
           
            # Process each page
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
               
                lines = text.split('\n')
               
                # Extract header information including month
                for i, line in enumerate(lines):
                    if 'ECR Of' in line or 'Contribution History' in line:
                        # Extract establishment code and period
                        match = re.search(r'(ECR Of|Contribution History.*?Of)\s+(\d+)\s+for\s+([A-Za-z]+\d+)', line)
                        if match:
                            extracted_data['header_info']['establishment_code'] = match.group(2)
                            extracted_data['header_info']['period'] = match.group(3)
                            # Extract month name
                            extracted_data['header_info']['month'] = extract_month_from_text(line)
                   
                    elif "Employees' State Insurance Corporation" in line:
                        extracted_data['header_info']['organization'] = line.strip()
                   
                    elif 'Total IP Contribution' in line and 'Total Employer Contribution' in line:
                        # Extract summary totals
                        next_line = lines[i + 1] if i + 1 < len(lines) else ""
                        amounts = re.findall(r'[\d,]+\.?\d*', next_line)
                        if len(amounts) >= 5:
                            extracted_data['summary_info'] = {
                                'total_ip_contribution': amounts[0],
                                'total_employer_contribution': amounts[1],
                                'total_contribution': amounts[2],
                                'total_government_contribution': amounts[3],
                                'total_monthly_wages': amounts[4]
                            }
               
                # Extract employee table data using text parsing approach
                employee_section_started = False
                employee_rows = []
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Check if this line contains employee data pattern
                    if re.search(r'^\d+\s+-\s+\d{10}', line):
                        employee_section_started = True
                        employee_rows.append(line)
                    elif employee_section_started and re.match(r'^\d+', line):
                        if not re.search(r'^\d+\s+-\s+\d{10}', line):
                            parts = line.split()
                            has_ip_pattern = False
                            for i, part in enumerate(parts):
                                if re.match(r'^\d{10}$', part) and i > 0:
                                    has_ip_pattern = True
                                    break
                            
                            if has_ip_pattern:
                                employee_rows.append(line)
                            else:
                                if employee_rows:
                                    employee_rows[-1] += ' ' + line
                    elif employee_section_started and line.lower().startswith(('page', 'printed')):
                        break
                
                # Process employee rows
                for row_text in employee_rows:
                    employee_record = parse_employee_row_improved(row_text, extracted_data['summary_info'], extracted_data['header_info'])
                    if employee_record:
                        extracted_data['employee_data'].append(employee_record)

                # Extract footer information
                if 'Printed On:' in text:
                    match = re.search(r'Printed On:\s*([^\n]+)', text)
                    if match:
                        extracted_data['footer_info']['printed_on'] = match.group(1).strip()
               
                if 'Page' in text:
                    match = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', text)
                    if match:
                        extracted_data['footer_info']['page_info'] = f"Page {match.group(1)} of {match.group(2)}"

        return extracted_data
   
    except Exception as e:
        st.error(f"Error extracting data: {str(e)}")
        return None


def parse_employee_row_improved(row_text, summary_info, header_info):
    """Parse individual employee row with improved logic for handling names and data"""
    try:
        row_text = row_text.strip()
        if not row_text:
            return None
        
        # Split the row into parts
        parts = row_text.split()
        if len(parts) < 6:  # Minimum required parts
            return None
        
        # Find the IP Number (10 digits) - this is our anchor
        ip_number = ""
        ip_index = -1
        
        for i, part in enumerate(parts):
            if re.match(r'^\d{10}$', part):
                ip_number = part
                ip_index = i
                break
        
        if not ip_number or ip_index < 2:  # Should have at least SNo and Is_Disable before IP
            return None
        
        # Extract SNo (first part, should be a number)
        sno = parts[0] if parts[0].isdigit() else "1"
        
        # Extract Is Disable (usually "-" and should be right before IP number)
        is_disable = "-"
        if ip_index > 0:
            is_disable = parts[ip_index - 1]
        
        # Everything after IP number until we hit numbers is the name
        name_parts = []
        data_parts = []
        
        # Start collecting name parts after IP number
        name_started = True
        for i in range(ip_index + 1, len(parts)):
            part = parts[i]
            
            # Check if this looks like numeric data (days, wages, contribution)
            if re.match(r'^\d+(\.\d{2})?$', part.replace(',', '')):
                # This is numeric data
                name_started = False
                data_parts.append(part.replace(',', ''))
            elif re.match(r'^(No|Work|Left|Service|Servic|-|Absent)$', part, re.IGNORECASE):
                # This is reason
                name_started = False
                data_parts.append(part)
            elif name_started:
                # This is part of the name
                name_parts.append(part)
            else:
                # We've started collecting data, but this doesn't look like data
                # This might be reason text
                data_parts.append(part)
        
        # Construct the name
        ip_name = ' '.join(name_parts).strip() if name_parts else "UNKNOWN"
        
        # Parse numeric data and reason
        days = "0"
        wages = "0.00"
        contribution = "0.00"
        reason = "-"
        
        # Separate numeric values from text values
        numeric_values = []
        text_values = []
        
        for part in data_parts:
            if re.match(r'^\d+(\.\d{2})?$', part):
                numeric_values.append(part)
            else:
                text_values.append(part)
        
        # Assign numeric values (usually in order: days, wages, contribution)
        if len(numeric_values) >= 1:
            if len(numeric_values) == 1:
                # Only one number, likely contribution
                contribution = numeric_values[0]
            elif len(numeric_values) == 2:
                # Two numbers, likely wages and contribution
                wages = numeric_values[0]
                contribution = numeric_values[1]
            elif len(numeric_values) >= 3:
                # Three or more numbers: days, wages, contribution
                days = numeric_values[0]
                wages = numeric_values[1]
                contribution = numeric_values[2]
        
        # Handle reason
        if text_values:
            reason_text = ' '.join(text_values).strip()
            if 'No' in reason_text and 'Work' in reason_text:
                reason = 'No Work'
            elif 'Left' in reason_text and ('Service' in reason_text or 'Servic' in reason_text):
                reason = 'Left Service'
            elif reason_text and reason_text != '-':
                reason = reason_text
            else:
                reason = '-'
        
        # Ensure proper decimal formatting for monetary values
        if '.' not in wages:
            wages += '.00'
        if '.' not in contribution:
            contribution += '.00'
        
        employee_record = {
            'SNo.': safe_numeric_convert(sno, is_integer=True),
            'Is Disable': is_disable,
            'IP Number': ip_number,  # Keep as string for IP numbers
            'IP Name': ip_name,
            'No. Of Days': safe_numeric_convert(days, is_integer=True),
            'Total Wages': safe_numeric_convert(wages),
            'IP Contribution': safe_numeric_convert(contribution),
            'Reason': reason,
            # Add month from header info
            'Month': header_info.get('month', 'Not Found'),
            # Add summary columns - convert to numbers
            'Total IP Contribution': safe_numeric_convert(summary_info.get('total_ip_contribution', '')),
            'Total Employer Contribution': safe_numeric_convert(summary_info.get('total_employer_contribution', '')),
            'Total Contribution': safe_numeric_convert(summary_info.get('total_contribution', '')),
            'Total Government Contribution': safe_numeric_convert(summary_info.get('total_government_contribution', '')),
            'Total Monthly Wages': safe_numeric_convert(summary_info.get('total_monthly_wages', ''))
        }
        
        return employee_record
        
    except Exception as e:
        st.error(f"Error parsing employee row: {row_text}, Error: {e}")
        return None


def format_excel_sheet(worksheet, data, start_row=1):
    """Apply formatting to Excel sheet to match PDF structure"""
    if not OPENPYXL_AVAILABLE:
        return start_row
   
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
   
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
   
    current_row = start_row
   
    # Add title/header information
    if 'header_info' in data:
        title = f"ECR Of {data['header_info'].get('establishment_code', '')} for {data['header_info'].get('period', '')}"
        worksheet.cell(row=current_row, column=1, value=title)
        worksheet.cell(row=current_row, column=1).font = Font(name='Arial', size=14, bold=True)
        worksheet.merge_cells(f'A{current_row}:I{current_row}')  # Updated to include month column
        current_row += 1
       
        org_name = data['header_info'].get('organization', '')
        if org_name:
            worksheet.cell(row=current_row, column=1, value=org_name)
            worksheet.cell(row=current_row, column=1).font = header_font
            worksheet.merge_cells(f'A{current_row}:I{current_row}')  # Updated to include month column
            current_row += 1
        
        # Add month information
        month_info = data['header_info'].get('month', '')
        if month_info and month_info != 'Not Found':
            worksheet.cell(row=current_row, column=1, value=f"Month: {month_info}")
            worksheet.cell(row=current_row, column=1).font = header_font
            worksheet.merge_cells(f'A{current_row}:I{current_row}')
            current_row += 1
   
    current_row += 1  # Add space
   
    # Add summary information
    if 'summary_info' in data:
        summary_headers = ['Total IP Contribution', 'Total Employer Contribution', 'Total Contribution',
                          'Total Government Contribution', 'Total Monthly Wages']
       
        for col, header in enumerate(summary_headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
       
        current_row += 1
       
        summary_values = [
            safe_numeric_convert(data['summary_info'].get('total_ip_contribution', '')),
            safe_numeric_convert(data['summary_info'].get('total_employer_contribution', '')),
            safe_numeric_convert(data['summary_info'].get('total_contribution', '')),
            safe_numeric_convert(data['summary_info'].get('total_government_contribution', '')),
            safe_numeric_convert(data['summary_info'].get('total_monthly_wages', ''))
        ]
       
        for col, value in enumerate(summary_values, 1):
            cell = worksheet.cell(row=current_row, column=col, value=value)
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = border
       
        current_row += 2  # Add space
   
    return current_row


def create_combined_excel(all_data):
    """Create single Excel file with all PDF data in separate sheets"""
    if not OPENPYXL_AVAILABLE:
        # Fallback to simple pandas Excel writer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Combine all employee data
            all_employee_data = []
            for file_data in all_data:
                filename = file_data['filename']
                data = file_data['data']
                
                if 'employee_data' in data and data['employee_data']:
                    for employee in data['employee_data']:
                        employee_copy = employee.copy()
                        employee_copy['Source_File'] = filename.replace('.pdf', '')
                        all_employee_data.append(employee_copy)
            
            if all_employee_data:
                df = pd.DataFrame(all_employee_data)
                df.to_excel(writer, sheet_name='Combined_Data', index=False)
        
        output.seek(0)
        return output
   
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create combined data sheet first
    combined_ws = wb.create_sheet("Combined_Data")
    
    # Combine all employee data
    all_employee_data = []
    for file_data in all_data:
        filename = file_data['filename']
        data = file_data['data']
        
        if 'employee_data' in data and data['employee_data']:
            for employee in data['employee_data']:
                employee_copy = employee.copy()
                employee_copy['Source_File'] = filename.replace('.pdf', '')
                all_employee_data.append(employee_copy)
    
    if all_employee_data:
        # Create combined data table - Updated headers to include month
        headers = ['Source_File', 'Month', 'SNo.', 'Is Disable', 'IP Number', 'IP Name', 'No. Of Days', 'Total Wages', 'IP Contribution', 'Reason']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = combined_ws.cell(row=1, column=col, value=header)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Write combined employee data
        for row_idx, employee in enumerate(all_employee_data, 2):
            for col, header in enumerate(headers, 1):
                value = employee.get(header, '')
                cell = combined_ws.cell(row=row_idx, column=col, value=value)
                cell.font = Font(name='Arial', size=9)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Center align numeric columns
                if any(keyword in header.lower() for keyword in ['contribution', 'wages', 'days', 'sno']):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Create individual sheets for each PDF
    for file_data in all_data:
        filename = file_data['filename']
        data = file_data['data']
        
        # Create sheet name (Excel sheet names have limitations)
        sheet_name = filename.replace('.pdf', '')[:31]  # Excel sheet name limit is 31 chars
        ws = wb.create_sheet(sheet_name)
        
        # Format the individual sheet
        current_row = format_excel_sheet(ws, data)
        
        # Add employee data table
        if 'employee_data' in data and data['employee_data']:
            # Updated headers to include month
            headers = ['Month', 'SNo.', 'Is Disable', 'IP Number', 'IP Name', 'No. Of Days', 'Total Wages', 'IP Contribution', 'Reason']
           
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
           
            current_row += 1
           
            # Write employee data
            for employee in data['employee_data']:
                for col, header in enumerate(headers, 1):
                    value = employee.get(header, '')
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.font = Font(name='Arial', size=9)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                   
                    # Center align numeric columns
                    if any(keyword in header.lower() for keyword in ['contribution', 'wages', 'days', 'sno']):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
               
                current_row += 1
        
        # Add footer information
        if 'footer_info' in data:
            current_row += 1
            if 'page_info' in data['footer_info']:
                ws.cell(row=current_row, column=1, value=data['footer_info']['page_info'])
                current_row += 1
           
            if 'printed_on' in data['footer_info']:
                ws.cell(row=current_row, column=1, value=f"Printed On: {data['footer_info']['printed_on']}")
        
        # Auto-adjust column widths for individual sheets
        from openpyxl.utils import get_column_letter
        
        for column_cells in ws.columns:
            max_length = 0
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)
           
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
           
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Auto-adjust column widths for combined sheet
    from openpyxl.utils import get_column_letter
    
    for column_cells in combined_ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)
       
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
       
        adjusted_width = min(max_length + 2, 50)
        combined_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================================
# ESIC CHALLAN EXTRACTOR
# ============================================================================

class ESICChallanExtractor:
    def __init__(self):
        self.required_keywords = [
            'esic', 'challan', 'employer', 'transaction', 'amount'
        ]
        
    def extract_text_pdfplumber(self, pdf_bytes):
        """Extract text using pdfplumber"""
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text
        except Exception as e:
            logger.error(f"Error extracting text with pdfplumber: {str(e)}")
            return None
    
    def extract_text_pymupdf(self, pdf_bytes):
        """Extract text using PyMuPDF"""
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            text = ""
            for page_num in range(doc.page_count):
                page = doc[page_num]
                text += page.get_text() + "\n"
            doc.close()
            return text
        except Exception as e:
            logger.error(f"Error extracting text with PyMuPDF: {str(e)}")
            return None
    
    def extract_text_from_pdf(self, pdf_bytes):
        """Extract text using available PDF library"""
        text = None
        
        if PDFPLUMBER_AVAILABLE:
            text = self.extract_text_pdfplumber(pdf_bytes)
        
        if text is None and PYMUPDF_AVAILABLE:
            text = self.extract_text_pymupdf(pdf_bytes)
        
        return text
    
    def check_esic_keywords(self, text):
        """Check if the PDF contains ESIC-related keywords"""
        if not text:
            return False
        
        text_lower = text.lower()
        found_keywords = sum(1 for keyword in self.required_keywords 
                           if keyword in text_lower)
        
        # Require at least 3 out of 5 keywords to be present
        return found_keywords >= 3
    
    def extract_field_patterns(self, text):
        """Extract specific fields using regex patterns"""
        patterns = {
            'transaction_status': [
                r'status[:\s]*([^\n\r]+)',
                r'transaction\s*status[:\s]*([^\n\r]+)',
                r'payment\s*status[:\s]*([^\n\r]+)'
            ],
            'employer_code': [
                r'employer[\'s\s]*code[:\s]*(\d+)',
                r'code\s*no[:\s]*(\d+)',
                r'employer\s*no[:\s]*(\d+)'
            ],
            'employer_name': [
                r'employer[\'s\s]*name[:\s]*([^\n\r]+)',
                r'name\s*of\s*employer[:\s]*([^\n\r]+)',
                r'establishment[:\s]*([^\n\r]+)'
            ],
            'challan_period': [
                r'challan\s*period[:\s]*([^\n\r]+)',
                r'period[:\s]*([^\n\r]+)',
                r'contribution\s*period[:\s]*([^\n\r]+)'
            ],
            'challan_number': [
                r'challan\s*no[:\s]*([A-Z0-9\-\/]+)',
                r'challan\s*number[:\s]*([A-Z0-9\-\/]+)',
                r'receipt\s*no[:\s]*([A-Z0-9\-\/]+)'
            ],
            'challan_created_date': [
                r'created\s*date[:\s]*(\d{1,2}[-\/]\d{1,2}[-\/]\d{4})',
                r'generation\s*date[:\s]*(\d{1,2}[-\/]\d{1,2}[-\/]\d{4})',
                r'date\s*of\s*creation[:\s]*(\d{1,2}[-\/]\d{1,2}[-\/]\d{4})'
            ],
            'challan_submitted_date': [
                r'submitted\s*date[:\s]*(\d{1,2}[-\/]\d{1,2}[-\/]\d{4})',
                r'payment\s*date[:\s]*(\d{1,2}[-\/]\d{1,2}[-\/]\d{4})',
                r'transaction\s*date[:\s]*(\d{1,2}[-\/]\d{1,2}[-\/]\d{4})'
            ],
            'amount_paid': [
                r'amount\s*paid[:\s]*₹?\s*([0-9,]+\.?\d*)',
                r'total\s*amount[:\s]*₹?\s*([0-9,]+\.?\d*)',
                r'paid\s*amount[:\s]*₹?\s*([0-9,]+\.?\d*)'
            ],
            'transaction_number': [
                # Enhanced patterns for transaction numbers
                r'transaction\s*(?:no|number|id)[:\s]*([A-Z0-9\-\/\.]+)',
                r'txn\s*(?:no|number|id)[:\s]*([A-Z0-9\-\/\.]+)',
                r'reference\s*(?:no|number|id)[:\s]*([A-Z0-9\-\/\.]+)',
                r'utr\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'bank\s*reference\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'payment\s*reference\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'ref\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'acknowledgment\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'ack\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'receipt\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                r'grn\s*(?:no|number)[:\s]*([A-Z0-9\-\/\.]+)',
                # Pattern for standalone alphanumeric codes (common in ESIC)
                r'(?:^|\n)\s*([A-Z]{2,}\d{6,}|\d{10,}[A-Z]+|\d{12,})\s*(?:\n|$)',
                # Pattern for transaction IDs in tables or structured format
                r'(?:transaction|txn|ref|reference)[\s\|]*([A-Z0-9]{8,})',
                # Additional loose patterns
                r'([A-Z0-9]{10,20})',  # Any alphanumeric string 10-20 chars
            ]
        }
        
        extracted_data = {}
        
        for field, pattern_list in patterns.items():
            value = None
            
            # Special handling for transaction_number with multiple attempts
            if field == 'transaction_number':
                value = self._extract_transaction_number(text, pattern_list)
            else:
                for pattern in pattern_list:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        value = match.group(1).strip()
                        break
            
            extracted_data[field] = value if value else "Not Found"
        
        return extracted_data
    
    def _extract_transaction_number(self, text, pattern_list):
        """Special method to extract transaction number with enhanced logic"""
        # First try the specific patterns
        for pattern in pattern_list[:-1]:  # Exclude the very loose pattern initially
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                candidate = match.group(1).strip()
                # Validate the candidate
                if self._is_valid_transaction_number(candidate):
                    return candidate
        
        # If no match found, try to find transaction numbers in common formats
        transaction_indicators = [
            'transaction', 'txn', 'reference', 'ref', 'utr', 'acknowledgment', 
            'ack', 'receipt', 'grn', 'bank', 'payment'
        ]
        
        # Look for lines containing transaction indicators
        lines = text.split('\n')
        for line in lines:
            line_lower = line.lower()
            for indicator in transaction_indicators:
                if indicator in line_lower:
                    # Extract potential transaction numbers from this line
                    potential_numbers = re.findall(r'[A-Z0-9]{8,20}', line, re.IGNORECASE)
                    for num in potential_numbers:
                        if self._is_valid_transaction_number(num):
                            return num
        
        # Last resort: look for any long alphanumeric strings
        all_codes = re.findall(r'\b[A-Z0-9]{10,20}\b', text, re.IGNORECASE)
        for code in all_codes:
            if self._is_valid_transaction_number(code):
                return code
        
        return None
    
    def _is_valid_transaction_number(self, candidate):
        """Validate if a candidate string looks like a valid transaction number"""
        if not candidate or len(candidate) < 8:
            return False
        
        # Remove common false positives
        false_positives = [
            'esic', 'challan', 'employer', 'employee', 'amount', 'total',
            'period', 'month', 'year', 'date', 'time', 'status', 'paid'
        ]
        
        if any(fp in candidate.lower() for fp in false_positives):
            return False
        
        # Check if it has a good mix of letters and numbers (typical for transaction IDs)
        has_letters = any(c.isalpha() for c in candidate)
        has_numbers = any(c.isdigit() for c in candidate)
        
        # Should have both letters and numbers, or be all numbers with good length
        if has_letters and has_numbers:
            return True
        elif has_numbers and not has_letters and len(candidate) >= 12:
            return True
        
        return False
    
    def extract_table_data(self, text):
        """Extract tabular data from the PDF"""
        tables = []
        
        # Look for table-like structures
        lines = text.split('\n')
        potential_table_lines = []
        
        for line in lines:
            # Check if line looks like a table row (has multiple columns separated by spaces/tabs)
            if re.search(r'\s+\d+\.\d{2}\s+|\s+₹\s*\d+', line) and len(line.split()) >= 3:
                potential_table_lines.append(line)
        
        if potential_table_lines:
            # Try to parse as table
            table_data = []
            for line in potential_table_lines:
                row = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
                if row:
                    table_data.append(row)
            tables.append(table_data)
        
        return tables
    
    def process_single_pdf(self, pdf_bytes, filename):
        """Process a single PDF file and extract ESIC challan data"""
        try:
            # Extract text
            text = self.extract_text_from_pdf(pdf_bytes)
            if not text:
                return {
                    'filename': filename,
                    'status': 'error',
                    'error': 'Could not extract text from PDF'
                }
            
            # Check if it's an ESIC document
            if not self.check_esic_keywords(text):
                return {
                    'filename': filename,
                    'status': 'not_esic',
                    'error': 'Document does not appear to be an ESIC challan'
                }
            
            # Extract structured data
            extracted_fields = self.extract_field_patterns(text)
            
            # Extract table data
            tables = self.extract_table_data(text)
            
            result = {
                'filename': filename,
                'status': 'success',
                'extracted_data': extracted_fields,
                'tables': tables,
                'raw_text': text[:1000] + "..." if len(text) > 1000 else text  # First 1000 chars
            }
            
            return result
            
        except Exception as e:
            logger.error(f"Error processing {filename}: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'filename': filename,
                'status': 'error',
                'error': str(e)
            }


def create_challan_excel_report(results):
    """Create Excel report from challan extraction results"""
    # Prepare data for DataFrame
    report_data = []
    
    for result in results:
        if result['status'] == 'success':
            extracted = result['extracted_data']
            row = {
                'Filename': result['filename'],
                'Status': result['status'],
                'Transaction Status': extracted.get('transaction_status', 'Not Found'),
                'Employer Code': extracted.get('employer_code', 'Not Found'),
                'Employer Name': extracted.get('employer_name', 'Not Found'),
                'Challan Period': extracted.get('challan_period', 'Not Found'),
                'Challan Number': extracted.get('challan_number', 'Not Found'),
                'Challan Created Date': extracted.get('challan_created_date', 'Not Found'),
                'Challan Submitted Date': extracted.get('challan_submitted_date', 'Not Found'),
                'Amount Paid': safe_numeric_convert_challan(extracted.get('amount_paid', 'Not Found')),
                'Transaction Number': extracted.get('transaction_number', 'Not Found'),
                'Tables Found': len(result.get('tables', [])),
                'Error': ''
            }
        else:
            row = {
                'Filename': result['filename'],
                'Status': result['status'],
                'Transaction Status': 'Error',
                'Employer Code': 'Error',
                'Employer Name': 'Error',
                'Challan Period': 'Error',
                'Challan Number': 'Error',
                'Challan Created Date': 'Error',
                'Challan Submitted Date': 'Error',
                'Amount Paid': 'Error',
                'Transaction Number': 'Error',
                'Tables Found': 0,
                'Error': result.get('error', 'Unknown error')
            }
        
        report_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(report_data)
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='ESIC_Challan_Report', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['ESIC_Challan_Report']
        
        # Format headers
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D7E4BD',
            'border': 1
        })
        
        # Format data cells
        cell_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'top',
            'border': 1
        })
        
        # Error cell format
        error_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'top',
            'border': 1,
            'fg_color': '#FFC7CE'
        })
        
        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        # Format data rows
        for row_num in range(1, len(df) + 1):
            for col_num, value in enumerate(df.iloc[row_num - 1]):
                if df.columns[col_num] == 'Status' and value in ['error', 'not_esic']:
                    worksheet.write(row_num, col_num, value, error_format)
                else:
                    worksheet.write(row_num, col_num, value, cell_format)
        
        # Auto-adjust column widths
        for i, column in enumerate(df.columns):
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(column)
            )
            worksheet.set_column(i, i, min(max_length + 2, 50))
    
    output.seek(0)
    return output

def create_enhanced_upload_section(title, description, key, file_type="pdf"):
    st.markdown(f"""
    <div class="upload-card animate-fadeInUp">
        <h3 style="color: #1e293b; font-family: 'Inter', sans-serif; font-weight: 600; margin-bottom: 1rem;">{title}</h3>
        <p style="color: #64748b; font-family: 'Inter', sans-serif; margin-bottom: 1.5rem;">{description}</p>
    </div>
    """, unsafe_allow_html=True)
    
    return st.file_uploader(
        f"Choose {file_type.upper()} files",
        type=file_type,
        accept_multiple_files=True,
        key=key,
        label_visibility="collapsed"
    )

# ============================================================================
# STREAMLIT APPLICATION
# ============================================================================

def main():
    st.set_page_config(
        page_title="ESIC PDF Data Extractor",
        page_icon="📄",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Enhanced Custom CSS for professional appearance
    st.markdown("""
    <style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Main container styling */
    .main > div {
        padding: 1rem 2rem;
    }
    
    /* Custom header styling */
    .custom-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 20px;
        margin-bottom: 2rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        position: relative;
        overflow: hidden;
    }
    
    .custom-header::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -50%;
        width: 200%;
        height: 200%;
        background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
        animation: shimmer 3s ease-in-out infinite;
    }
    
    @keyframes shimmer {
        0%, 100% { transform: translateX(-100%) translateY(-100%) rotate(45deg); }
        50% { transform: translateX(-50%) translateY(-50%) rotate(45deg); }
    }
    
    .header-content {
    display: flex;
    align-items: center;
    position: relative;
    z-index: 2;
    min-height: 200px; /* Add minimum height to accommodate larger logo */
    }

    .logo-section img {
    border-radius: 15px;
    box-shadow: 0 5px 15px rgba(0,0,0,0.2);
    transition: transform 0.3s ease;
    max-width: 100%; /* Ensure responsiveness */
    height: auto;
    }

/* Add responsive adjustments */
@media (max-width: 768px) {
    .logo-section img {
        width: 300px !important; /* Smaller on mobile */
    }
}
    
    .logo-section img {
        border-radius: 15px;
        box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        transition: transform 0.3s ease;
    }
    
    .logo-section img:hover {
        transform: scale(1.05);
    }
    
    .title-section {
        flex-grow: 1;
    }
    
    .main-title {
        color: white !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 700 !important;
        font-size: 3rem !important;
        margin: 0 !important;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        letter-spacing: -0.02em;
    }
    
    .subtitle {
        color: rgba(255,255,255,0.9) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 1.2rem !important;
        margin: 0.5rem 0 0 0 !important;
        font-weight: 300 !important;
    }
    
    /* Enhanced tabs styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
        background-color: #f8fafc;
        padding: 0.5rem;
        border-radius: 15px;
        border: none;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05);
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 60px;
        white-space: pre-wrap;
        background-color: transparent;
        border-radius: 10px;
        color: #64748b;
        font-weight: 500;
        font-family: 'Inter', sans-serif;
        border: none;
        padding: 0 2rem;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%);
        color: white !important;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background-color: rgba(59, 130, 246, 0.1);
        transform: translateY(-2px);
    }
    
    /* Enhanced cards and containers */
    .upload-card {
        background: white;
        padding: 2rem;
        border-radius: 20px;
        box-shadow: 0 5px 25px rgba(0,0,0,0.08);
        border: 1px solid #e2e8f0;
        margin: 1rem 0;
        transition: all 0.3s ease;
    }
    
    .upload-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 15px 35px rgba(0,0,0,0.15);
    }
    
    .stats-card {
        background: linear-gradient(135deg, #f1f5f9 0%, #e2e8f0 100%);
        padding: 1.5rem;
        border-radius: 15px;
        border: 1px solid #cbd5e1;
        margin: 0.5rem 0;
        text-align: center;
        transition: all 0.3s ease;
    }
    
    .stats-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.1);
    }
    
    /* Enhanced buttons */
    .stButton > button {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        border-radius: 12px;
        font-weight: 600;
        font-family: 'Inter', sans-serif;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(16, 185, 129, 0.3);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(16, 185, 129, 0.4);
        background: linear-gradient(135deg, #059669 0%, #047857 100%);
    }
    
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%);
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);
    }
    
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1d4ed8 0%, #1e40af 100%);
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.4);
    }
    
    /* Enhanced metrics */
    [data-testid="metric-container"] {
        background: white;
        padding: 1.5rem;
        border-radius: 15px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05);
        transition: all 0.3s ease;
    }
    
    [data-testid="metric-container"]:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.1);
    }
    
    [data-testid="metric-container"] [data-testid="metric-container-label"] {
        font-family: 'Inter', sans-serif;
        font-weight: 500;
        color: #64748b;
    }
    
    [data-testid="metric-container"] [data-testid="metric-container-value"] {
        font-family: 'Inter', sans-serif;
        font-weight: 700;
        font-size: 2rem;
        color: #1e293b;
    }
    
    /* Enhanced file uploader */
    .stFileUploader > div {
        background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
        border: 2px dashed #cbd5e1;
        border-radius: 20px;
        padding: 3rem;
        text-align: center;
        transition: all 0.3s ease;
    }
    
    .stFileUploader > div:hover {
        border-color: #3b82f6;
        background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
        transform: translateY(-2px);
    }
    
    /* Enhanced dataframes */
    .stDataFrame {
        border-radius: 15px;
        overflow: hidden;
        box-shadow: 0 5px 25px rgba(0,0,0,0.08);
    }
    
    /* Progress bars */
    .stProgress .st-bo {
        background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%);
        border-radius: 10px;
    }
    
    /* Enhanced expanders */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
        border-radius: 15px;
        font-family: 'Inter', sans-serif;
        font-weight: 600;
    }
    
    /* Footer enhancement */
    .footer-enhanced {
        background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
        color: white;
        padding: 2rem;
        border-radius: 20px;
        margin-top: 3rem;
        text-align: center;
        box-shadow: 0 5px 25px rgba(0,0,0,0.1);
    }
    
    .footer-enhanced p {
        margin: 0.5rem 0;
        font-family: 'Inter', sans-serif;
    }
    
    /* Animations */
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .animate-fadeInUp {
        animation: fadeInUp 0.6s ease-out;
    }
    
    /* Responsive design */
    @media (max-width: 768px) {
        .header-content {
            flex-direction: column;
            text-align: center;
        }
        
        .logo-section {
            margin-right: 0;
            margin-bottom: 1rem;
        }
        
        .main-title {
            font-size: 2rem !important;
        }
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Enhanced header with professional design
    # Enhanced header with professional design and proper logo placement
    try:
        from PIL import Image
        logo = Image.open("kkc logo.png")
        
        st.markdown('<div class="custom-header">', unsafe_allow_html=True)

        # Create columns for logo and title
        col_logo, col_title = st.columns([3, 7])
        
        with col_logo:
            st.image(logo, width=700)  # Set to 750px width as requested
        
        with col_title:
            st.markdown("""
                <div style="padding-left: 2rem; display: flex; flex-direction: column; justify-content: center; height: 150px;">
                    <h1 style="color: black; font-family: 'Inter', sans-serif; font-weight: 700; font-size: 2.5rem; margin: 0; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">
                        ESIC PDF Data Extractor
                    </h1>
                    <p style="color: rgba(255,255,255,0.9); font-family: 'Inter', sans-serif; font-size: 1.1rem; margin: 0.5rem 0 0 0; font-weight: 300;">
                        Professional • Efficient • Reliable Data Processing
                    </p>
                </div>
            """, unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
    except Exception as e:
        # Fallback header without logo
        st.markdown("""
        <div class="custom-header">
            <div style="text-align: center; padding: 2rem;">
                <h1 style="color: white; font-family: 'Inter', sans-serif; font-weight: 700; font-size: 3rem; margin: 0; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">
                    ESIC PDF Data Extractor
                </h1>
                <p style="color: rgba(255,255,255,0.9); font-family: 'Inter', sans-serif; font-size: 1.2rem; margin: 0.5rem 0 0 0; font-weight: 300;">
                    Professional • Efficient • Reliable Data Processing
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.warning(f"Logo not loaded: {str(e)}")
    
    # Rest of your existing code continues here...    
    # Check for required libraries
    missing_libraries = []
    if not PDFPLUMBER_AVAILABLE:
        missing_libraries.append("pdfplumber")
    if not PYMUPDF_AVAILABLE:
        missing_libraries.append("PyMuPDF (fitz)")
    if not OPENPYXL_AVAILABLE:
        missing_libraries.append("openpyxl")
    
    if missing_libraries:
        st.warning(f"⚠️ Missing libraries: {', '.join(missing_libraries)}. Install them for full functionality.")
    
    # Create tabs for different functionalities
    tab1, tab2 = st.tabs(["📊 ECR Extractor", "💰 Challan Extractor"])
    
    # ============================================================================
    # TAB 1: CONTRIBUTION HISTORY EXTRACTOR
    # ============================================================================
    with tab1:
        
        if not PDFPLUMBER_AVAILABLE:
            st.error("❌ pdfplumber is required for contribution history extraction. Please install it first.")
            st.code("pip install pdfplumber")
            return
        
        uploaded_files = create_enhanced_upload_section(
            "ESIC ECR File Upload", 
            "Upload ESIC ECR PDF files to extract employee contribution data including month information.",
            "contribution_files"
        )
        
        if uploaded_files:
            st.info(f"📁 Selected {len(uploaded_files)} file(s) for processing")
            
            if st.button("🔄 Process ECR PDFs", type="primary", key="process_contribution"):
                # Create containers for different sections
                progress_container = st.container()
                results_container = st.container()
                download_container = st.container()
                
                with progress_container:
                    st.subheader("🔄 Processing Status")
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Processing results tracking
                    all_data = []
                    successful_files = []
                    failed_files = []
                    
                    # Process files
                    for i, uploaded_file in enumerate(uploaded_files):
                        status_text.text(f"Processing: {uploaded_file.name}")
                        progress_bar.progress((i + 1) / len(uploaded_files))
                        
                        try:
                            extracted_data = extract_esic_data(uploaded_file)
                            if extracted_data:
                                all_data.append({
                                    'filename': uploaded_file.name,
                                    'data': extracted_data
                                })
                                successful_files.append(uploaded_file.name)
                            else:
                                failed_files.append(uploaded_file.name)
                        
                        except Exception as e:
                            failed_files.append(f"{uploaded_file.name} (Error: {str(e)})")
                    
                    status_text.empty()
                    progress_bar.empty()
                
                # Show results summary
                with results_container:
                    if all_data or failed_files:
                        st.subheader("📊 Processing Summary")
                        
                        # Summary metrics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        total_employees = sum(len(data['data'].get('employee_data', [])) for data in all_data)
                        total_files = len(uploaded_files)
                        successful_count = len(successful_files)
                        failed_count = len(failed_files)
                        
                        with col1:
                            st.metric("📄 Total Files", total_files)
                        with col2:
                            st.metric("✅ Successful", successful_count, delta=f"{(successful_count/total_files*100):.1f}%")
                        with col3:
                            st.metric("❌ Failed", failed_count, delta=f"{(failed_count/total_files*100):.1f}%" if failed_count > 0 else "0%")
                        with col4:
                            st.metric("👥 Total Employees", total_employees)
                        
                        # Success/failure indicator
                        if successful_count == total_files:
                            st.success(f"🎉 All {total_files} files processed successfully!")
                        elif successful_count > 0:
                            st.warning(f"⚠️ {successful_count} files processed successfully, {failed_count} failed")
                        else:
                            st.error("❌ No files were processed successfully")
                
                # Download section and preview
                with download_container:
                    if all_data:
                        st.subheader("📥 Download & Preview")
                        
                        col1, col2 = st.columns([2, 1])
                        
                        with col1:
                            # Generate Excel file
                            try:
                                excel_file = create_combined_excel(all_data)
                                
                                st.download_button(
                                    label="📥 Download Excel Report",
                                    data=excel_file,
                                    file_name=f"ESIC_ECR_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                                
                            except Exception as e:
                                st.error(f"❌ Error creating Excel file: {str(e)}")
                        
                        with col2:
                            st.info(f"💡 Excel contains:\n• Combined data sheet with month info\n• Individual file sheets\n• {total_employees} employee records")
                        
                        # Data preview
                        if all_data[0]['data'].get('employee_data'):
                            st.subheader("📋 Data Preview (First 10 rows)")
                            preview_df = pd.DataFrame(all_data[0]['data']['employee_data'][:10])
                            # Show only key columns for preview including month
                            key_columns = ['Month', 'SNo.', 'IP Number', 'IP Name', 'No. Of Days', 'Total Wages', 'IP Contribution']
                            available_columns = [col for col in key_columns if col in preview_df.columns]
                            if available_columns:
                                st.dataframe(preview_df[available_columns], use_container_width=True)
                    
                    # Show processing details in collapsible section
                    if successful_files or failed_files:
                        with st.expander("📝 View Processing Details", expanded=False):
                            if successful_files:
                                st.success("✅ Successfully Processed Files:")
                                for filename in successful_files:
                                    # Show extracted month if available
                                    file_data = next((item for item in all_data if item['filename'] == filename), None)
                                    if file_data and 'header_info' in file_data['data']:
                                        month = file_data['data']['header_info'].get('month', 'Unknown')
                                        st.write(f"• {filename} (Month: {month})")
                                    else:
                                        st.write(f"• {filename}")
                            
                            if failed_files:
                                st.error("❌ Failed Files:")
                                for filename in failed_files:
                                    st.write(f"• {filename}")
    
    # ============================================================================
    # TAB 2: CHALLAN EXTRACTOR
    # ============================================================================
    with tab2:
        
        if not PDFPLUMBER_AVAILABLE and not PYMUPDF_AVAILABLE:
            st.error("❌ Either pdfplumber or PyMuPDF is required for challan extraction.")
            st.code("pip install pdfplumber PyMuPDF")
            return
        
        uploaded_challan_files = create_enhanced_upload_section(
            "ESIC Challan File Upload", 
            "Upload ESIC challan PDF files for transaction and payment data extraction.",
            "challan_files"
        )
        
        if uploaded_challan_files:
            st.info(f"📁 Selected {len(uploaded_challan_files)} file(s) for processing")
            
            if st.button("🔄 Process Challan PDFs", type="primary", key="process_challan"):
                # Create containers
                progress_container = st.container()
                results_container = st.container()
                download_container = st.container()
                
                with progress_container:
                    st.subheader("🔄 Processing Status")
                    extractor = ESICChallanExtractor()
                    results = []
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    for i, uploaded_file in enumerate(uploaded_challan_files):
                        status_text.text(f"Processing: {uploaded_file.name}")
                        progress_bar.progress((i + 1) / len(uploaded_challan_files))
                        
                        pdf_bytes = uploaded_file.read()
                        result = extractor.process_single_pdf(pdf_bytes, uploaded_file.name)
                        results.append(result)
                    
                    status_text.empty()
                    progress_bar.empty()
                
                # Results summary
                with results_container:
                    if results:
                        st.subheader("📊 Processing Summary")
                        
                        # Calculate statistics
                        successful = sum(1 for r in results if r['status'] == 'success')
                        failed = sum(1 for r in results if r['status'] == 'error')
                        not_esic = sum(1 for r in results if r['status'] == 'not_esic')
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("📄 Total Files", len(results))
                        with col2:
                            st.metric("✅ Successful", successful, delta=f"{(successful/len(results)*100):.1f}%")
                        with col3:
                            st.metric("❌ Failed", failed, delta=f"{(failed/len(results)*100):.1f}%" if failed > 0 else "0%")
                        with col4:
                            st.metric("⚠️ Not ESIC", not_esic, delta=f"{(not_esic/len(results)*100):.1f}%" if not_esic > 0 else "0%")
                        
                        # Status indicator
                        if successful == len(results):
                            st.success(f"🎉 All {len(results)} files processed successfully!")
                        elif successful > 0:
                            st.warning(f"⚠️ {successful} files processed successfully, {failed + not_esic} had issues")
                        else:
                            st.error("❌ No files were processed successfully")
                
                # Download and preview section
                with download_container:
                    if results:
                        st.subheader("📥 Download & Preview")
                        
                        col1, col2 = st.columns([2, 1])
                        
                        with col1:
                            # Generate Excel report
                            try:
                                excel_report = create_challan_excel_report(results)
                                
                                st.download_button(
                                    label="📥 Download Challan Data",
                                    data=excel_report,
                                    file_name=f"ESIC_Challan_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                                
                            except Exception as e:
                                st.error(f"❌ Error creating Excel report: {str(e)}")
                        
                        with col2:
                            st.info(f"💡 Report contains:\n• All file processing results\n• Extracted field data\n• Error details")
                        
                        # Quick preview of successful extractions
                        successful_results = [r for r in results if r['status'] == 'success']
                        if successful_results:
                            st.subheader("📋 Quick Preview - Successfully Extracted Data")
                            
                            preview_data = []
                            for result in successful_results[:5]:  # Show first 5 successful results
                                data = result['extracted_data']
                                preview_data.append({
                                    'Filename': result['filename'][:30] + "..." if len(result['filename']) > 30 else result['filename'],
                                    'Transaction Status': data.get('transaction_status', 'N/A')[:20],
                                    'Employer Code': data.get('employer_code', 'N/A'),
                                    'Amount Paid': data.get('amount_paid', 'N/A'),
                                    'Transaction Number': data.get('transaction_number', 'N/A')[:15] + "..." if len(str(data.get('transaction_number', 'N/A'))) > 15 else data.get('transaction_number', 'N/A')
                                })
                            
                            if preview_data:
                                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
                        
                        # Detailed results in collapsible section
                        with st.expander("📝 View Detailed Extraction Results", expanded=False):
                            for result in results:
                                status_icon = "✅" if result['status'] == 'success' else "❌" if result['status'] == 'error' else "⚠️"
                                
                                with st.container():
                                    st.markdown(f"**{status_icon} {result['filename']}**")
                                    
                                    if result['status'] == 'success':
                                        data = result['extracted_data']
                                        
                                        col1, col2 = st.columns(2)
                                        with col1:
                                            st.write("**Transaction Details:**")
                                            st.write(f"• Status: {data.get('transaction_status', 'N/A')}")
                                            st.write(f"• Transaction Number: {data.get('transaction_number', 'N/A')}")
                                            st.write(f"• Amount Paid: {data.get('amount_paid', 'N/A')}")
                                        
                                        with col2:
                                            st.write("**Employer Details:**")
                                            st.write(f"• Employer Code: {data.get('employer_code', 'N/A')}")
                                            st.write(f"• Challan Period: {data.get('challan_period', 'N/A')}")
                                            st.write(f"• Tables Found: {len(result.get('tables', []))}")
                                    
                                    else:
                                        st.error(f"Error: {result.get('error', 'Unknown error')}")
                                    
                                    st.markdown("---")

    # ============================================================================
    # FOOTER
    # ============================================================================
    st.markdown("""
    <div class="footer-enhanced">
        <h3 style="margin-bottom: 1rem; font-weight: 600;">🔧 ESIC PDF Data Extractor</h3>
        <p style="color: rgba(255,255,255,0.8);">📧 For issues or feature requests, please contact the development team</p>
        <div style="margin-top: 1rem; padding-top: 1rem; border-top: 1px solid rgba(255,255,255,0.2);">
            <p style="font-size: 0.9rem; color: rgba(255,255,255,0.6);">🔄 Refresh to restart from the beginning</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
